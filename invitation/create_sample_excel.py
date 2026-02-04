#!/usr/bin/env python3
"""
Create sample Excel file with recipient data
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# Create workbook
wb = Workbook()
ws = wb.active
ws.title = "Recipients"

# Style header row
header_fill = PatternFill(start_color="667EEA", end_color="667EEA", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=12)

# Add headers
ws['A1'] = "Name"
ws['B1'] = "Email"

ws['A1'].fill = header_fill
ws['A1'].font = header_font
ws['A1'].alignment = Alignment(horizontal="center", vertical="center")

ws['B1'].fill = header_fill
ws['B1'].font = header_font
ws['B1'].alignment = Alignment(horizontal="center", vertical="center")

# Set column widths
ws.column_dimensions['A'].width = 25
ws.column_dimensions['B'].width = 35

# Add sample data
sample_data = [
    ["Rajesh Kumar", "rajesh.kumar@example.com"],
    ["Priya Sharma", "priya.sharma@example.com"],
    ["Amit Patel", "amit.patel@example.com"],
    ["Sneha Reddy", "sneha.reddy@example.com"],
    ["Vikram Singh", "vikram.singh@example.com"],
]

for idx, (name, email) in enumerate(sample_data, start=2):
    ws[f'A{idx}'] = name
    ws[f'B{idx}'] = email

# Save file
wb.save("recipients.xlsx")
print("✅ Created recipients.xlsx with sample data")
print("\n📝 Sample recipients:")
for name, email in sample_data:
    print(f"   - {name} ({email})")
print("\n⚠️  Replace with actual recipient data before sending!")
