#!/usr/bin/env python3
"""

Reads recipient details from Excel and sends HTML-formatted invitations via SMTP
"""

import smtplib
import os
import sys
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.image import MIMEImage
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook
from dotenv import load_dotenv
from concurrent.futures import ThreadPoolExecutor, as_completed
import threading
import time
from PIL import Image, ImageDraw, ImageFont
import shutil

# Load environment variables from .env file if it exists
load_dotenv()

SCRIPT_DIR = Path(__file__).resolve().parent


def load_logos_for_email():
    """
    Load logo files and return dict of {cid: (bytes, mime_subtype)}.
    Logos are attached as inline MIME parts and referenced via cid: in HTML.
    """
    logos = {}
    configs = [
        ('sm_logo', 'sm_logo_small.png', 'png'),
    ]
    for cid, filename, subtype in configs:
        path = SCRIPT_DIR / filename
        try:
            with open(path, 'rb') as f:
                logos[cid] = (f.read(), subtype)
        except FileNotFoundError:
            pass
    return logos


def get_html_template(name="Volunteer"):
    """Generate a premium HTML email template with WhatsApp group integration."""

    return f"""
<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<title>🎉 Congratulations – Welcome to SM</title>
</head>

<body style="margin:0;padding:0;background:#0f2027;font-family:'Segoe UI',Arial,sans-serif;">

<!-- PREVIEW TEXT (INBOX LINE) -->
<div style="display:none;max-height:0;overflow:hidden;">
🎉 Congratulations {name}! Welcome to the SM Volunteers Forum
</div>

<!-- HERO BLAST -->
<table width="100%" cellpadding="0" cellspacing="0">
  <tr>
    <td align="center" style="
      background:linear-gradient(135deg,#0f2027,#203a43,#2c5364);
      padding:80px 15px;
      color:#ffffff;
    ">
      <h1 style="
        font-size:42px;
        margin:0;
        letter-spacing:2px;
        text-transform:uppercase;
      ">
        🎉 CONGRATULATIONS 🎉
      </h1>

      <p style="font-size:18px;margin-top:15px;">
        Welcome to the SM Volunteers Forum
      </p>
    </td>
  </tr>
</table>

<!-- CONTENT CARD -->
<table width="100%" cellpadding="0" cellspacing="0" style="background:#f4f6f8;">
  <tr>
    <td align="center" style="padding:40px 15px;">

      <table width="100%" cellpadding="0" cellspacing="0"
        style="
          max-width:620px;
          background:#ffffff;
          border-radius:18px;
          box-shadow:0 18px 45px rgba(0,0,0,0.25);
        ">

        <!-- SM LOGO -->
        <tr>
          <td align="center" style="padding-top:35px;">
            <img src="sm_logo.png"
                 alt="SM Logo"
                 width="90"
                 style="display:block;">
          </td>
        </tr>

        <!-- BODY -->
        <tr>
          <td style="padding:35px 40px;color:#222222;">

            <p style="font-size:16px;">
              Dear <strong>{name}</strong>,
            </p>

            <p style="font-size:16px;line-height:1.8;">
              We are delighted to inform you that you have
              <strong>successfully cleared the SM interview</strong>.
              Your dedication, confidence, and commitment truly stood out.
            </p>

            <p style="font-size:16px;line-height:1.8;">
              You are now an <strong>official member of the SM Team</strong> and
              eligible to participate in all
              <strong>SM events, initiatives, and activities</strong>.
            </p>

            <!-- HIGHLIGHT -->
            <div style="
              margin:30px 0;
              padding:22px;
              background:linear-gradient(135deg,#e3f2fd,#fce4ec);
              border-radius:14px;
              text-align:center;
              font-weight:bold;
              font-size:15px;
            ">
              ✨ Welcome to the SM Family ✨
            </div>

            <!-- WHATSAPP INFO -->
            <p style="font-size:16px;font-weight:bold;">
              📲 Kindly Join the Official SM WhatsApp Group
            </p>

            <p style="font-size:14px;line-height:1.7;">
              All upcoming information, announcements, and updates
              will be shared <strong>only through the official WhatsApp group</strong>.
              Kindly ensure that you join the group to stay informed.
            </p>

            <div style="text-align:center;margin:35px 0;">
              <a href="https://chat.whatsapp.com/CJeFwL5abHc8VkqeAa3n1v"
                 style="
                   background:linear-gradient(135deg,#25D366,#1ebe5d);
                   color:#ffffff;
                   padding:16px 38px;
                   text-decoration:none;
                   font-size:16px;
                   font-weight:bold;
                   border-radius:50px;
                   display:inline-block;
                 ">
                🚀 Join WhatsApp Group
              </a>
            </div>

            <p style="font-size:13px;color:#666;">
              ⚠️ Joining the WhatsApp group is <strong>mandatory</strong>.
            </p>

            <p style="margin-top:35px;font-size:15px;">
              Achievements are earned through dedication and hard work.
              Congratulations on this proud milestone!
            </p>

            <p style="margin-top:20px;">
              Warm regards,<br>
              <strong>SM Team</strong>
            </p>

          </td>
        </tr>

      </table>

    </td>
  </tr>
</table>

</body>
</html>
"""

def read_recipients_from_excel(file_path):
    """Read recipient details from Excel file (Email and optional Name)"""
    recipients = []
    
    try:
        workbook = load_workbook(filename=file_path, read_only=True)
        sheet = workbook.active
        
        # Get headers
        header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
        email_idx = 0
        name_idx = -1
        
        # Try to find columns by header names
        for i, val in enumerate(header_row):
            if val:
                val_lower = str(val).lower()
                if 'email' in val_lower:
                    email_idx = i
                elif 'name' in val_lower:
                    name_idx = i
        
        # Skip header row and read data
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if len(row) > email_idx and row[email_idx]:
                email = str(row[email_idx]).strip()
                name = "Volunteer"
                if name_idx != -1 and len(row) > name_idx and row[name_idx]:
                    name = str(row[name_idx]).strip()
                
                if email:
                    recipients.append({'email': email, 'name': name})
        
        workbook.close()
        return recipients
    
    except FileNotFoundError:
        print(f"❌ Error: Excel file not found at {file_path}")
        return None
    except Exception as e:
        print(f"❌ Error reading Excel file: {str(e)}")
        return None


def get_smtp_config():
    """Get SMTP configuration from environment variables or user input"""
    config = {}
    
    # Try to get from environment variables first
    config['server'] = "smtp.gmail.com"
    config['port'] = "587"
    config['email'] = "smvolunteers@ksrct.ac.in"  # Change to your Gmail address
    config['password'] = "mxod vqth nulr ywoh"  # gmail app password
    
    # If not in environment, prompt user
    if not all([config['server'], config['port'], config['email'], config['password']]):
        print("\n📧 SMTP Configuration")
        print("=" * 50)
        
        if not config['server']:
            print("\nCommon SMTP servers:")
            print("  Gmail: smtp.gmail.com")
            print("  Outlook: smtp-mail.outlook.com")
            config['server'] = input("Enter SMTP server: ").strip()
        
        if not config['port']:
            config['port'] = input("Enter SMTP port (default 587): ").strip() or "587"
        
        if not config['email']:
            config['email'] = input("Enter your email address: ").strip()
        
        if not config['password']:
            import getpass
            config['password'] = getpass.getpass("Enter your email password/app password: ")
    
    config['port'] = int(config['port'])
    return config


def generate_invitation_image(name, template_path="Congratulations.png", output_dir="generated_invites"):
    """
    Generate a personalized image with the recipient's name drawn on the template.
    Returns the path to the generated image file.
    """
    try:
        # Create output directory if it doesn't exist
        os.makedirs(output_dir, exist_ok=True)
        
        # Sanitize filename
        safe_name = "".join([c for c in name if c.isalnum() or c in (' ', '_', '-')]).strip()
        if not safe_name:
            safe_name = "Guest"
        output_filename = f"Invitation_{safe_name}.png"
        output_path = os.path.join(output_dir, output_filename)
        
        # Check if template exists
        if not os.path.exists(template_path):
            print(f"⚠️ Template '{template_path}' not found! Skipping image generation.")
            return None
            
        img = Image.open(template_path)
        draw = ImageDraw.Draw(img)
        width, height = img.size

        # --- PATCHING LOGIC TO REMOVE "NAME" ---
        # Strategy: Tile a clean section of the ribbon to cover the center area.
        # This prevents the blurry look of stretching.
        
        y_ribbon_start = int(height * 0.69)
        y_ribbon_end = int(height * 0.77)
        ribbon_height = y_ribbon_end - y_ribbon_start
        
        # Target area to cover (Center where "NAME" is)
        target_x_start = int(width * 0.38) # Narrowed slightly to be safe
        target_x_end = int(width * 0.62)
        target_width = target_x_end - target_x_start
        
        # Source area (Clean ribbon on left)
        src_x_start = int(width * 0.20)
        src_x_end = int(width * 0.28) # Take a smaller, safer clean chunk
        src_width = src_x_end - src_x_start
        
        clean_slice = img.crop((src_x_start, y_ribbon_start, src_x_end, y_ribbon_end))
        
        # Tile the slice to cover the target width
        current_x = target_x_start
        while current_x < target_x_end:
            paste_width = min(src_width, target_x_end - current_x)
            if paste_width < src_width:
                # Crop the last piece if needed
                patch = clean_slice.crop((0, 0, paste_width, ribbon_height))
            else:
                patch = clean_slice
            
            img.paste(patch, (current_x, y_ribbon_start))
            current_x += src_width
        
        # --- DRAWING TEXT ---
        
        name = name.upper() # FORCE UPPERCASE
        
        # Color: Dark Maroon
        text_color = (60, 0, 0) 
        
        # Max width for text 
        max_text_width = int(width * 0.55) 
        
        # Dynamic Scaler
        current_font_size = int(width * 0.06) # Start bigger
        min_font_size = int(width * 0.03)
        
        # Load Font - Switch to Serif (Times New Roman) for premium look
        font_names = ["timesbd.ttf", "georgiab.ttf", "arialbd.ttf"]
        font_path = None
        for fn in font_names:
            possible_path = f"C:/Windows/Fonts/{fn}"
            if os.path.exists(possible_path):
                font_path = possible_path
                break
        
        if not font_path:
             font_path = "C:/Windows/Fonts/arial.ttf"

        while current_font_size > min_font_size:
            try:
                font = ImageFont.truetype(font_path, current_font_size)
            except OSError:
                font = ImageFont.load_default()
                break

            if hasattr(draw, "textbbox"):
                bbox = draw.textbbox((0, 0), name, font=font)
                text_w = bbox[2] - bbox[0]
                text_h = bbox[3] - bbox[1]
            else:
                text_w, text_h = draw.textsize(name, font=font)
                
            if text_w <= max_text_width:
                break 
            
            current_font_size -= 2
        
        # Calculate centered position
        x = (width - text_w) / 2
        
        # Center vertically in the ribbon patch
        ribbon_middle = y_ribbon_start + (ribbon_height / 2)
        # Adjust vertical center slightly for font baseline
        y = ribbon_middle - (text_h / 2) - (text_h * 0.15) 

        # Draw text
        draw.text((x, y), name, font=font, fill=text_color)

        img.save(output_path)
        return output_path

    except Exception as e:
        print(f"❌ Error generating image for {name}: {str(e)}")
        return None


def send_single_email(recipient, smtp_config, logos, idx, total):
    """Send a single email to one recipient (thread-safe)."""
    email = recipient['email']
    name = recipient['name'] 
    
    try:
        # Create new connection for this thread
        server = smtplib.SMTP(smtp_config['server'], smtp_config['port'], timeout=90)
        server.starttls()
        server.login(smtp_config['email'], smtp_config['password'])
        
        # Prepare HTML Content using CSS to mimic the design (No image generation)
        # Colors picked from the original design
        
        html_content = f"""
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Congratulations</title>
</head>
<body style="margin: 0; padding: 0; background-color: #f4f4f4; font-family: 'Arial', sans-serif;">
    
    <table width="100%" cellpadding="0" cellspacing="0" style="background-color: #f4f4f4; padding: 40px 0;">
        <tr>
            <td align="center">
                
                <!-- Main Card (Restored with Border & Background) -->
                <div style="
                    max-width: 600px; 
                    margin: 0 auto; 
                    background-color: #420000; /* Card Background */
                    border: 2px solid #ffd700; /* Gold Border */
                    border-radius: 15px;
                    padding: 40px 20px; 
                    box-shadow: 0 10px 25px rgba(0,0,0,0.5);
                    text-align: center;
                    color: #ffffff;
                ">
                    
                    <!-- Top Gold Decoration Line -->
                    <div style="height: 2px; background: #ffd700; margin-bottom: 30px;"></div>

                    <!-- Congratulations Text -->
                    <h1 style="
                        font-family: 'Times New Roman', Times, serif; 
                        color: #ffffff; 
                        font-size: 36px; 
                        margin: 0 0 10px 0; 
                        font-style: italic;
                        font-weight: normal;
                        text-shadow: 0 2px 4px rgba(0,0,0,0.5);
                    ">
                        Congratulations
                    </h1>

                    <h2 style="
                        font-family: 'Arial', sans-serif;
                        color: #ffc107; 
                        font-size: 20px; 
                        margin: 0 0 35px 0; 
                        text-transform: uppercase; 
                        letter-spacing: 2px;
                        line-height: 1.4;
                    ">
                        WELCOME TO OUR<br>
                        <span style="color: #ffd700; font-size: 24px; font-weight: bold;">SM VOLUNTEERS FORUM</span>
                    </h2>

                    <!-- Logo (Fixed Round Shape with Container) -->
                    <div style="margin: 0 auto 25px auto; width: 160px; height: 160px; border-radius: 50%; border: 3px solid #b8860b; overflow: hidden; background-color: transparent;">
                        <img src="cid:sm_logo" alt="SM Volunteers" style="display: block; width: 100%; height: 100%; object-fit: cover;">
                    </div>

                    <!-- Name Display (Transparent with Gold Border) -->
                    <div style="
                        background-color: transparent;
                        border: 2px solid #ffd700;
                        padding: 15px 0;
                        margin: 30px auto;
                        width: 80%;
                        border-radius: 8px;
                        box-shadow: 0 0 15px rgba(255, 215, 0, 0.1);
                    ">
                        <span style="
                            font-family: 'Arial Black', 'Arial Bold', sans-serif;
                            font-size: 30px;
                            color: #ffd700; /* Gold Text */
                            text-transform: uppercase;
                            letter-spacing: 2px;
                            display: block;
                            font-weight: 900;
                            text-shadow: 0 2px 4px rgba(0,0,0,0.5);
                        ">
                            {name.upper()}
                        </span>
                    </div>

                    <!-- Main Body Content -->
                    <div style="text-align: left; padding: 0 20px; color: #e0e0e0; font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; font-size: 16px; line-height: 1.6;">
                        <p>
                            We are delighted to inform you that you have <strong>successfully cleared the SM interview</strong>. 
                            Your dedication, confidence, and commitment truly stood out.
                        </p>
                        <p>
                            You are now an <strong>official member of the SM Team</strong> and eligible to participate in all 
                            <strong>SM events, initiatives, and activities</strong>.
                        </p>
                        <p style="margin-top: 25px; font-size: 14px; color: #cccccc;">
                            All upcoming information, announcements, and updates will be shared 
                            <strong>only through the official WhatsApp group</strong>. 
                            Kindly ensure that you join the group to stay informed.
                        </p>
                    </div>

                    <!-- Footer Text -->
                    <p style="
                        font-family: 'Georgia', serif;
                        font-size: 15px;
                        color: #cccccc;
                        line-height: 1.6;
                        margin-top: 35px;
                        font-style: italic;
                        padding: 0 20px;
                    ">
                        Achievements are earned through dedication and hard work.<br>
                        Congratulations on this proud milestone!
                    </p>
                    
                    <!-- WhatsApp Button (Solid Green) -->
                    <div style="margin-top: 40px; margin-bottom: 10px;">
                        <a href="https://chat.whatsapp.com/CJeFwL5abHc8VkqeAa3n1v" style="
                            background-color: #25D366; /* Solid Green */
                            color: white;
                            padding: 12px 30px;
                            text-decoration: none;
                            border-radius: 25px;
                            font-family: sans-serif;
                            font-weight: bold;
                            box-shadow: 0 4px 10px rgba(37, 211, 102, 0.3);
                            display: inline-block;
                        ">
                            Join WhatsApp Group
                        </a>
                    </div>

                </div>
                
                <p style="color: #666666; font-size: 11px; margin-top: 20px; font-family: sans-serif;">
                    KSRCT SM Volunteers
                </p>

            </td>
        </tr>
    </table>
</body>
</html>
        """
        
        # Determine logos to attach
        # We need 'sm_logo' for the HTML cid:sm_logo
        
        msg = MIMEMultipart('related')
        msg['Subject'] = f"Congratulations {name}! - SM Volunteers"
        msg['From'] = f"SM Official <{smtp_config['email']}>"
        msg['To'] = email

        html_part = MIMEText(html_content, 'html', 'utf-8')
        msg.attach(html_part)
        
        # Attach Inline Logos
        # We assume load_logos_for_email returns 'sm_logo' key
        if logos:
             for cid, (payload, subtype) in logos.items():
                img = MIMEImage(payload, _subtype=subtype)
                img.add_header('Content-Disposition', 'inline', filename=cid)
                img.add_header('Content-ID', f'<{cid}>')
                msg.attach(img)
        else:
            print("⚠️ Warning: Logos not found. Email will be missing images.")

        print(f"✅ [{idx}/{total}] Prepared HTML email for {name}")
        
        # Send email
        server.send_message(msg)
        server.quit()
        
        print(f"🚀 [{idx}/{total}] Sent to {email}")
        return {'status': 'success', 'email': email}
        
    except Exception as e:
        print(f"❌ [{idx}/{total}] Failed to send to {email}: {str(e)}")
        return {'status': 'failed', 'email': email, 'error': str(e)}


def send_invitation_emails(recipients, smtp_config, excel_file):
    """Send invitation emails to all recipients with inline logo images (CID) using parallel processing."""
    
    print(f"\n📨 Preparing to send {len(recipients)} invitation emails...")
    print("=" * 50)
    
    # Test connection first
    print(f"\n🔌 Testing connection to {smtp_config['server']}:{smtp_config['port']}...")
    try:
        server = smtplib.SMTP(smtp_config['server'], smtp_config['port'], timeout=30)
        server.starttls()
        server.login(smtp_config['email'], smtp_config['password'])
        server.quit()
        print("✅ Connection successful!\n")
    except smtplib.SMTPAuthenticationError:
        print("\n❌ Authentication failed! Please check your email and password.")
        print("   For Gmail, make sure you're using an App Password, not your regular password.")
        print("   Generate one at: https://myaccount.google.com/apppasswords")
        return None, None
    except Exception as e:
        print(f"\n❌ Error connecting to SMTP server: {str(e)}")
        return None, None
    
    logos = load_logos_for_email()
    if logos:
        print(f"🖼️  Logos loaded: {', '.join(logos.keys())}")
    else:
        print("⚠️  No logo files found; images will not display inline.")
    
    successful = []
    failed = []
    
    # Send emails sequentially with delay to avoid spam filters
    print("📬 Sending emails sequentially (slow mode to avoid spam)...\n")
    for idx, recipient in enumerate(recipients, 1):
        result = send_single_email(recipient, smtp_config, logos, idx, len(recipients))
        if result['status'] == 'success':
            successful.append(result['email'])
        else:
            failed.append({'email': result['email'], 'error': result.get('error', 'Unknown error')})
        time.sleep(2)  # 2-second delay between emails
    
    print("\n" + "=" * 50)
    
    # Print summary
    print(f"\n📊 Summary:")
    print(f"   ✅ Successfully sent: {len(successful)}")
    print(f"   ❌ Failed: {len(failed)}")
    
    return successful, failed


def main():
    """Main function"""
    print("\n" + "=" * 60)
    print("  SM Volunteers - Official Selection Notifier")
    print("  K. S. Rangasamy College of Technology")
    print("=" * 60)
    
    # Use fixed Excel file path
    excel_file = "recipients.xlsx"
    
    # Check if file exists
    if not Path(excel_file).exists():
        print(f"\n❌ Error: File '{excel_file}' not found!")
        print("\nPlease make sure the Excel file exists with columns:")
        print("  - Column A: Email")
        sys.exit(1)
    
    # Read recipients
    print(f"\n📖 Reading recipients from {excel_file}...")
    recipients = read_recipients_from_excel(excel_file)
    
    if not recipients:
        print("\n❌ No valid recipients found in the Excel file!")
        sys.exit(1)
    
    print(f"✅ Found {len(recipients)} recipients")
    
    # Show preview
    print("\n📋 Preview of recipients:")
    for i, r in enumerate(recipients[:5], 1):
        print(f"   {i}. {r['name']} <{r['email']}>")
    if len(recipients) > 5:
        print(f"   ... and {len(recipients) - 5} more")
    
    # Confirm
    confirm = input("\n⚠️  Proceed with sending emails? (yes/no): ").strip().lower()
    if confirm not in ['yes', 'y']:
        print("\n❌ Cancelled by user")
        sys.exit(0)
    
    # Get SMTP config
    smtp_config = get_smtp_config()
    
    # Send emails
    successful, failed = send_invitation_emails(recipients, smtp_config, excel_file)
    
    if successful is not None:
        print("\n✅ Email sending process completed!")
    else:
        print("\n❌ Email sending process failed!")
        sys.exit(1)


if __name__ == "__main__":
    main()
