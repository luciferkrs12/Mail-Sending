#!/usr/bin/env python3
"""
Add real recipient data to recipients.xlsx
"""
from openpyxl import load_workbook
from openpyxl.styles import Alignment

def add_recipients():
    file_name = "recipients.xlsx"
    try:
        wb = load_workbook(file_name)
        ws = wb.active
        
        # Clear existing data (but keep headers)
        if ws.max_row > 1:
            ws.delete_rows(2, ws.max_row)
        
        # New recipient data
        new_data = [
            ["Mohit Raj", "mohitur669@gmail.com"],
        ]
        
        for idx, (name, email) in enumerate(new_data, start=2):
            ws[f'A{idx}'] = name
            ws[f'B{idx}'] = email
            ws[f'A{idx}'].alignment = Alignment(horizontal="left")
            ws[f'B{idx}'].alignment = Alignment(horizontal="left")
            
        wb.save(file_name)
        print(f"✅ Successfully updated {file_name} with {len(new_data)} recipients.")
        for name, email in new_data:
            print(f"   👤 {name} ({email})")
            
    except Exception as e:
        print(f"❌ Error: {e}")

if __name__ == '__main__':
    add_recipients()
