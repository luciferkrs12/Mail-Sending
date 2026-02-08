#!/usr/bin/env python3
"""
Update Excel file with Name and Email columns for SM Volunteers
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

def create_template():
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "SM_Volunteers"

    # Style header row
    # Using a nice Royal Blue to match the email theme
    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)

    # Add headers
    ws['A1'] = "Name"
    ws['B1'] = "Email"

    # Apply styles to headers
    for cell in ['A1', 'B1']:
        ws[cell].fill = header_fill
        ws[cell].font = header_font
        ws[cell].alignment = Alignment(horizontal="center", vertical="center")

    # Set column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 40

    # Add sample data
    sample_data = [
        ["Mohit Raj", "mohitur669@gmail.com"],
        ["Sample Name", "sample.email@example.com"],
    ]

    for idx, (name, email) in enumerate(sample_data, start=2):
        ws[f'A{idx}'] = name
        ws[f'B{idx}'] = email
        ws[f'A{idx}'].alignment = Alignment(horizontal="left")
        ws[f'B{idx}'].alignment = Alignment(horizontal="left")

    # Save file - overwriting previous recipients.xlsx
    file_name = "recipients.xlsx"
    wb.save(file_name)
    print(f"✅ Created {file_name} with Name and Email columns.")
    print("📝 Headers: Name | Email")

if __name__ == '__main__':
    create_template()
